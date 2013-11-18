#!/usr/bin/python
# To change what data is shown in the excelWorkbook for each trace type, edit the function: tools.classes.rObject.makeSpecific()
import openpyxl, argparse, os
from pyrecon.tools.classes import loadSeries, rObject
from operator import attrgetter

def main(path_to_series, save_path):
    if save_path[-1] != '/':
        save_path += '/'
    if not os.path.exists(save_path):
        print 'Creating new directory: '+save_path
        os.mkdir(save_path)
    if '.xlsx' not in save_path:
        save_path += path_to_series.replace('.ser','').split('/')[-1]
        save_path += '.xlsx'
    series = loadSeries(path_to_series)
    wkbk = Workbook(series=series)
    wkbk.getDendrites()
    wkbk.getProtrusions()
    wkbk.listProtrusionChildren()
    wkbk.writeProtrusionsPerDendrite()
    wkbk.writeProtrusionChildrenToProtrusions()        
    wkbk.save(save_path)

class Workbook(openpyxl.Workbook):
    def __init__(self, series=None):
        openpyxl.Workbook.__init__(self)
        
        self.series = series # Series object for this workbook
        self.objects = series.getObjectLists() # Names of objects in this series
        self.filterType = ['c'] # Ignore these rTypes
    def listProtrusionChildren(self): #===
        childList = []
        for protrusion in self.protrusions:
            for child in protrusion.children:
                childList.append(child)
        self.protrusionChildren = sorted(list(set(childList)))
    def getProtrusions(self):
        '''Gathering protrusions/children from series'''
        protrusions = []
        for protName in self.objects[1]:
            protrusion = rObject(name=protName, series=self.series)
            protrusions.append(protrusion)
        self.protrusions = protrusions
    def getDendrites(self):
        dendrites = []
        for dendName in self.objects[0]:
            dendrite = rObject(name=dendName, series=self.series)
            dendrites.append(dendrite)
        self.dendrites = dendrites
    def writeProtrusionsPerDendrite(self):
        '''Creates a column of protrusions, in order by start section.
        All other data objects are aligned with their protrusion's row'''

        # Create worksheet for each dendrite
        for dendrite in self.dendrites:
            row = 1
            column = 0
            self.create_sheet(title=dendrite.name)
            sheet = self.get_sheet_by_name(dendrite.name)
            
            # Create a row for each protrusion in that dendrite
            for protrusion in sorted(self.protrusions, key=attrgetter('start','name')): #Sort by start # (by prot# if same start#)
                column = 0
                # Check if needs to be in this dendrite sheet
                if protrusion.dendrite == dendrite.name:
                    sheet.cell(row=row, column=column).value = protrusion.name
                    column+=1
                    # Write prot data
                    for data_item in protrusion.data:
                        sheet.cell(row=0, column=column).value = str(data_item)
                        sheet.cell(row=row, column=column).value = protrusion.data[data_item]
                        column+=1
                    row += 1+protrusion.getSpacing()                
    
    def writeProtrusionChildrenToProtrusions(self): #=== make more concise
        # Grab existing wrksht for each dendrite
        for dendrite in self.dendrites:
            sheet = self.get_sheet_by_name(dendrite.name)
            
            # Get list of protrusions in this dendrite
#             protList = [prot for prot in self.protrusions if dendrite.name in prot.name]
            protList = [prot for prot in self.protrusions if dendrite.name == prot.dendrite] #===
            # Get list of all protrusion children in this dendrite
            childList = []
            for prot in protList:
                for child in prot.children:
                    childList.append(child)
            childList = sorted(list(set(childList))) # Unique child names

            # Go through childList
            row = 1
            column = 5
            for child in childList:
                # find prots with this child
                for prot in protList:
                    if child in prot.children:
                        
                        # update row to correct protrusion row
                        for row in range(sheet.get_highest_row()):
                            if sheet.cell(row=row, column=0).value == prot.name:
                                row = row
                                break
                            
                        # Add data for each subChild 
                        subColumn = column
                        for subChild in prot.children[child]:
                            subColumn = column
                            subChildObj = rObject(name=subChild, series=self.series)
                            sheet.cell(row=row, column=subColumn).value = subChildObj.name
                            subColumn = column+1
                            for data_item in subChildObj.data:
                                sheet.cell(row=0, column=subColumn).value = str(data_item)
                                sheet.cell(row=row, column=subColumn).value = subChildObj.data[data_item]
                                subColumn+=1
                            row+=1
                column=subColumn+1
                        
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Creates an excel workbook containing protrusions and data, YAY!')
    parser.add_argument('series', nargs=1, type=str, help='Path to the series/sections that needs to be re-scaled')
    parser.add_argument('savepath', nargs=1, help='Directory where the excel workbook will be saved')
    args = vars(parser.parse_args())
    # Assign argparse things to their variables
    path_to_series = str(args['series'][0])
    save_path = str(args['savepath'][0])
    main(path_to_series, save_path)